from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import os
from datetime import date
from config import TEMPLATE_DIR, OUTPUT_DIR
from calculate_billing import get_billing_period


# =========================
# MAIN ENTRY
# =========================

def generate_reports(results, data):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for result in results:
        generate_single_report(result, data)


# =========================
# SINGLE REPORT
# =========================

def generate_single_report(result, data):
    template_path = os.path.join(TEMPLATE_DIR, "billing_template.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    tenant = get_tenant(result["tenant_id"], data)
    building = data["building"]
    unit = get_unit(result["unit_id"], data)
    year = data["year"]
    is_shop = unit.get("is_shop")
    remarks = {"*": False,
               "**": False}

    # =========================
    # HEADER (FIXED CELLS)
    # =========================

    # sender
    ws["F1"] = building.get("sender")
    ws["F2"] = building.get("sender_address") or ""

    # bank details
    ws["F5"] = building.get("owner")
    ws["F6"] = f"IBAN: {building.get("bank_account")}"
    ws["F7"] = building.get("bank") or ""

    # address
    ws["A7"] = build_recipients_name(tenant, 1)
    ws["A8"] = build_recipients_name(tenant, 2)
    ws["A9"] = building.get("address")
    ws["A10"] = building.get("city_line")
    if tenant.get("new_address"):
        ws["A9"] = tenant.get("new_address")
        ws["A10"] = tenant.get("new_city_line")

    # general information
    ws["F10"] = "Wohn- und Geschäftshaus" if building.get("has_shops", False) else "Wohnhaus"
    ws["F11"] = f"{building.get("address")}, {building.get("city_line")}"
    period = get_billing_period(building, year)
    ws["F12"] = (f"Abrechnungszeitraum: "
                 f"{period[0].strftime("%d.%m.%y")} - {period[1].strftime("%d.%m.%y")}")

    # information for calculation
    ws["H20"] = tenant.get("prepay_ops") or 0
    ws["H19"] = result.get("months")
    row = 18
    if not building["is_single_unit"]:
        ws[f"F{row}"] = "Ihre Wohnfläche:" if not is_shop else "Ihre Nutzfläche:"
        ws[f"H{row}"] = unit.get("area")
        ws[f"I{row}"] = "qm"
        row -= 1
        ws[f"F{row}"] = "Ihre Lage:" if not is_shop else "Ihre Lage:"
        ws[f"H{row}"] = unit.get("position")
        row -= 1

        if ((building.get("gar_count") or 0) > 0) and tenant.get("gar_count") > 0:
            ws[f"F{row}"] = "Ihre Garagen:"
            ws[f"H{row}"] = tenant.get("gar_count")
            row -= 1
            ws[f"F{row}"] = "Anzahl Garagen:"
            ws[f"H{row}"] = building.get("gar_count")
            row -= 1

        if (building.get("unit_count") or 0) > 0:
            ws[f"F{row}"] = "Anzahl Wohnungen:"
            ws[f"H{row}"] = building.get("unit_count")
            row -= 1

        # if there are costs that explicitly exclude shops
        if (building.get("total_tenant_area") or 0) > 0 and not tenant.get("is_shop"):
            ws[f"F{row}"] = "Gesamtwohnfl.*:"
            ws[f"H{row}"] = building.get("total_tenant_area")
            ws[f"I{row}"] = "qm"
            row -= 1
            remarks["*"] = True

        if (building.get("total_area") or 0) > 0:
            ws[f"F{row}"] = "Gesamtwohnnutzfl.:" if (building.get("has_shops") or False) else "Gesamtwohnfläche:"
            ws[f"H{row}"] = building.get("total_area")
            ws[f"I{row}"] = "qm"
            row -= 1

    # =========================
    # COST TABLE
    # =========================

    # header
    header_row = 22
    if building["is_single_unit"]:
        ws[f"I{header_row}"] = "Kosten (€)"
    else:
        ws[f"E{header_row}"] = "Gesamtkosten (€)"
        ws[f"G{header_row}"] = "Verteilt"
        ws[f"I{header_row}"] = "Ihr Anteil (€)"

    # costs
    row = header_row + 2
    total_amount_sum = 0
    for line in result["lines"]:
        if line["amount"] > 0:
            ws[f"A{row}"] = line.get("cost_type")
            ws[f"I{row}"] = line.get("amount")

            if not building["is_single_unit"]:
                ws[f"E{row}"] = line.get("total_amount") or ""
                ws[f"G{row}"] = line.get("allocation") or ""
                total_amount_sum += line.get("total_amount") or 0

                # special case: usage
                if line.get("type") == "individual" and (line.get("usage") or 0) > 0:
                    ws[f"C{header_row}"] = "Verbrauch x Preis"
                    ws[f"B{row}"] = f"{line.get("usage")} cbm x {round(line.get("price"), 4)}  €/cbm"

                # special case: usage of shops removed before
                if line.get("type") == "general" and (line.get("special_amount") or 0) > 0:
                    ws[f"C{header_row}"] = "* (€)"
                    ws[f"C{row}"] = line.get("special_amount")
                    # adjust name slightly
                    ws[f"A{row}"] = line.get("cost_type").removesuffix(" +")

                # special case: people
                if line.get("allocation") or "" == "Personen**":
                    remarks["**"] = True

            row += 1

    # total costs

    row -= 1
    draw_bottom_border(ws, row, ["I"])
    row += 2

    ws[f"A{row}"] = "Gesamtkosten"
    if not building["is_single_unit"]:
        ws[f"E{row}"] = total_amount_sum
    ws[f"E{row}"] = result["total_tenant_cost"]
    row += 1

    prepayment = tenant.get("prepay_ops") * result.get("months")
    ws[f"A{row}"] = "Ihre Vorauszahlung"
    ws[f"I{row}"] = prepayment
    draw_bottom_border(ws, row, ["I"])
    row += 2

    # =========================
    # FINAL BALANCE
    # =========================

    balance = prepayment - result["total_tenant_cost"]

    ws[f"A{row}"] = "Guthaben" if balance >= 0 else "Nachzahlung"
    ws[f"I{row}"] = abs(balance)

    # add tax
    if is_shop and building.get("has_tax"):
        tax = round(abs(balance) * 0.19, 2)
        ws[f"A{row}"] = "Guthaben Netto" if balance >= 0 else "Nachzahlung Netto"
        row += 1
        ws[f"A{row}"] = "zzgl. 19 % MwSt."
        ws[f"I{row}"] = tax
        row += 1
        draw_bottom_border(ws, row, ["I"])
        ws[f"A{row}"] = "Guthaben Brutto" if balance >= 0 else "Nachzahlung Brutto"
        ws[f"I{row}"] = abs(balance) + tax

    # apply formatting
    apply_template_fill(ws, source_row=header_row, target_row=row)
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"I{row}"].font = Font(bold=True)
    row += 3

    # =========================
    # ADDITIONAL REMARKS
    # =========================

    if balance >= 0:
        ws[f"A{row}"] = "Das Guthaben wird Ihnen per Banküberweisung erstattet."
    else:
        ws[f"A{row}"] = ("Die Nachzahlung überweisen Sie bitte mit der nächsten Mietzahlung "
                         "auf das oben genannte Bankkonto.")
    row += 2

    if not is_shop:
        ws[f"A{row}"] = "Etwaige Hausmeister- und Schornsteinfegerkosten enthalten ausschließlich Arbeitsleistungen "
        row += 1
        ws[f"A{row}"] = "und sind gemäß § 35a EStG steuerlich begünstigt."
        row += 2

    if remarks["*"]:
        ws[f"A{row}"] = "* ohne Gewerbe"
        row += 1

    if remarks["**"]:
        ws[f"A{row}"] = "** Berechnung der Anteile:"
        row += 1
        ws[f"A{row}"] = "Lage"
        ws[f"E{row}"] = "Personen"
        ws[f"F{row}"] = "x"
        ws[f"G{row}"] = "Monate"
        ws[f"H{row}"] = "Anteile"
        row += 1

        people_map = result.get("maps").get("people")
        occupancy_map = result.get("maps").get("occupancy")
        for tenant_id in people_map:
            cur_tenant = get_tenant(tenant_id, data)
            cur_unit =  get_unit(cur_tenant.get("unit_id"), data)
            ws[f"A{row}"] = cur_unit.get("position")
            ws[f"E{row}"] = people_map.get(tenant_id)
            ws[f"G{row}"] = occupancy_map.get(tenant_id)
            ws[f"H{row}"] = people_map.get(tenant_id) * occupancy_map.get(tenant_id)
            row += 1

        draw_bottom_border(ws, row - 1, ["H"])
        ws[f"H{row}"] = result.get("maps").get("special").get("people")

        row += 2

    # =========================
    # END OF LETTER
    # =========================

    ws[f"A{row}"] = "Gauting, den"
    ws[f"B{row}"] = date.today()
    ws[f"G{row}"] = building.get("sender")

    # =========================
    # SAVE FILE
    # =========================

    filename = f"{OUTPUT_DIR}/Betriebskosten_{result['tenant_id']}_{year}.xlsx"
    wb.save(filename)


# =========================
# HELPERS
# =========================

def draw_bottom_border(ws, row, cols):
    border = Border(bottom=Side(style="thin"))

    for col in cols:
        ws[f"{col}{row}"].border = border


def apply_template_fill(ws, source_row, target_row):
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws[f"{col}{target_row}"].fill = ws[f"{col}{source_row}"].fill


def get_tenant(tenant_id, data):
    for t in data["tenants"]:
        if t["tenant_id"] == tenant_id:
            return t
    raise ValueError("Tenant not found")


def get_unit(unit_id, data):
    for u in data["units"]:
        if u["unit_id"] == unit_id:
            return u
    raise ValueError("Unit not found")


def build_recipients_name(tenant, n):
    name = tenant.get(f"salutation{n}", "").strip()
    if name == "Herr":
        name = "Herrn"
    if tenant.get(f"title{n}").strip() is not None:
        name += f" {tenant.get(f"title{n}").strip()}"
    if tenant.get(f"first_name{n}").strip() is not None:
        name += f" {tenant.get(f"first_name{n}").strip()}"
    if tenant.get(f"last_name{n}").strip() is not None:
        name += f" {tenant.get(f"last_name{n}").strip()}"

    return name
